
import xlwt
# 1. 替换xlrd为openpyxl读取XLSX文件
from openpyxl import load_workbook

# 2. 使用绝对路径并修复文件读取方式
# wb_for_read = xlrd.open_workbook(r'Day21-30\res\Day24\数据\2020年销售数据.xlsx')
wb_for_read = load_workbook(r'D:\学习python100day计划\Day21-30\res\Day24\数据\2020年销售数据.xlsx')
# sheet1 = wb_for_read.sheet_by_index(0)
sheet1 = wb_for_read.active
nrows, ncols = sheet1.max_row, sheet1.max_column

# 3. 创建新的写工作簿并复制数据（替代xlutils.copy）
# wb_for_write = copy(wb_for_read)
# sheet2 = wb_for_write.get_sheet(0)
wb_for_write = xlwt.Workbook()
sheet2 = wb_for_write.add_sheet('Sheet1')

# 复制原始数据（openpyxl是1-based索引，xlwt是0-based）
for row in range(1, nrows + 1):
    for col in range(1, ncols + 1):
        cell_value = sheet1.cell(row=row, column=col).value
        sheet2.write(row - 1, col - 1, cell_value)

# 4. 修正公式范围（确保从数据行开始计算）
sheet2.write(nrows, 4, xlwt.Formula(f'AVERAGE(E2:E{nrows})'))  # E列平均值
sheet2.write(nrows, 6, xlwt.Formula(f'SUM(G2:G{nrows})'))      # G列总和

# 5. 保存文件
wb_for_write.save('阿里巴巴2020年股票数据汇总.xls')

