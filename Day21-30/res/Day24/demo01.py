from openpyxl import load_workbook

wb = load_workbook(r'D:\学习python100day计划\Day21-30\res\Day24\数据\2022年股票数据.xlsx')

# 获取工作簿中所有工作表的名称列表
sheetnames = wb.sheetnames
# 打印工作表名称列表，确认当前Excel文件包含的工作表
print(sheetnames)

# 选择第一个工作表进行数据处理
sheet = wb[sheetnames[0]]
# 打印当前工作表的最大行数和列数，用于确认数据范围
print(sheet.max_row, sheet.max_column)

# 遍历工作表中的所有行（row从0开始计数，对应Excel中的1-based行号）
for row in range(sheet.max_row):
    # 遍历当前行中的所有列（col从0开始计数，对应Excel中的1-based列号）
    for col in range(sheet.max_column):
        # 获取单元格值（openpyxl使用1-based索引，因此需要row+1和col+1）
        value = sheet.cell(row+1, col+1).value
        
        # 对数据行（排除表头行，row>0）进行格式化处理
        if row > 0:
            # 第1列（col=0）为日期类型，格式化为"YYYY年MM月DD日"形式
            if col == 0:
                value = f'{value.year}年{value.month:>02d}月{value.day:>02d}日'
            # 其他列（数值类型）格式化为带两位小数的千分位格式
            else:
                value = f'{value:,.2f}'
        
        # 打印当前单元格值，使用制表符分隔，不换行
        print(value, end='	')
    # 当前行所有列打印完成后换行
    print()

# 获取最后一个单元格的数据类型（n:数字, s:字符串, d:日期等）
last_cell_type = sheet.cell(sheet.max_row, sheet.max_column).data_type
# 打印最后一个单元格的数据类型
print(last_cell_type)

# 打印第2行所有单元格的值（sheet[2]表示Excel中的第2行，1-based索引）
print([cell.value for cell in sheet[2]])

# 打印第4行从第1列到第5列的单元格值（[0:5]表示前5列，0-based索引）
print([cell.value for cell in sheet[4][0:5]])
