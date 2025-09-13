# 导入openpyxl库，用于创建和操作Excel文件
import openpyxl
# 导入random库，用于生成随机成绩
import random

# 创建一个新的Excel工作簿
wb = openpyxl.Workbook()
# 获取当前活动工作表
sheet = wb.active
# 设置工作表标题为'成绩单'
sheet.title = '成绩单'

# 定义表头数据，包含姓名和各科成绩列
titles = ['姓名', '语文', '数学', '英语', '总分']
# 遍历表头列表，将表头写入Excel第一行（行索引从1开始）
for col_index, title in enumerate(titles):
    # cell(row, column, value)：行号、列号、单元格值
    sheet.cell(1, col_index + 1, title)

# 定义学生姓名列表
names = ['张三', '李四', '王五', '赵六', '钱七']
# 遍历学生姓名列表，为每个学生生成成绩并计算总分
for row_index, name in enumerate(names):
    # 将学生姓名写入第1列（A列），行号从2开始（row_index+2）
    sheet.cell(row_index + 2, 1, name)
    # 初始化总分变量
    total_score = 0
    # 遍历2-4列（B-D列），对应语文、数学、英语科目
    for col_index in range(2, 5):
        # 生成50-100之间的随机整数作为成绩
        score = random.randrange(50, 101)
        # 将成绩写入当前学生行的对应科目列
        sheet.cell(row_index + 2, col_index, score)
        # 将当前科目的成绩累加到总分
        total_score += score
    # 将计算好的总分写入第5列（E列）
    sheet.cell(row_index + 2, 5, total_score)

# 保存Excel文件到当前目录，文件名为'成绩单.xlsx'
wb.save(r'成绩单.xlsx')
