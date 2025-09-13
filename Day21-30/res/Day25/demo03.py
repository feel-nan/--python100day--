# 导入openpyxl库用于操作Excel文件
import openpyxl
# 从openpyxl.styles模块导入样式相关类，用于设置单元格格式
from openpyxl.styles import Font, Alignment, Border, Side

# 创建对齐方式对象：水平居中对齐，垂直居中对齐
alignment = Alignment(horizontal='center', vertical='center')

# 创建边框样式对象：中等虚线，黑色（RGB颜色代码'000000'）
side = Side(style='mediumDashed', color='000000')

# 加载已存在的Excel文件'成绩单.xlsx'
wb = openpyxl.load_workbook('成绩单.xlsx')
# 获取工作簿中的第一个工作表
sheet = wb.worksheets[0]

# 设置第1行（表头行）的高度为30
sheet.row_dimensions[1].height = 30
# 设置E列（平均分列）的宽度为120
sheet.column_dimensions['E'].width = 120

# 在E1单元格（第1行第5列）写入表头文本'平均分'
sheet['E1'] = '平均分'

# 设置E1单元格的字体样式：18号字，粗体，粉色（RGB颜色代码'ff1493'），华文楷体
sheet.cell(1, 5).font = Font(size=18, bold=True, color='ff1493', name='华文楷体')

# 应用之前定义的居中对齐方式到E1单元格
sheet.cell(1, 5).alignment = alignment

# 为E1单元格添加边框：四个方向都应用之前定义的虚线边框
sheet.cell(1, 5).border = Border(left=side, right=side, top=side, bottom=side)

# 遍历2到6行（学生数据行）
for i in range(2,7):
    # 在Ei单元格写入平均分计算公式：计算B列到D列（语文到英语）的平均值
    sheet[f'E{i}'] = f'=AVERAGE(B{i}:D{i})'
    # 设置当前平均分单元格的字体样式：14号字，粗体，粉色，华文楷体
    sheet.cell(i, 5).font = Font(size=14, bold=True, color='ff1493', name='华文楷体')
    # 应用居中对齐方式到当前平均分单元格
sheet.cell(i, 5).alignment = alignment
    # 为当前平均分单元格添加边框
sheet.cell(i, 5).border = Border(left=side, right=side, top=side, bottom=side)

# 保存修改后的Excel文件
wb.save('成绩单.xlsx')
